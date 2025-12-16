"""
Carta Cap Table Transformer
Transforms Carta exports into firm's internal cap table format.
Can be called standalone or via xlwings from Excel.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
from datetime import datetime
import re
from copy import copy


def parse_carta_export(carta_path: str) -> dict:
    """Parse Carta export and extract relevant data."""

    # Read raw to find header row (look for 'Stakeholder ID' or 'Name')
    raw = pd.read_excel(carta_path, sheet_name='Detailed Cap', header=None)

    header_row = None
    for i in range(min(10, len(raw))):
        row_values = [str(v).lower() for v in raw.iloc[i].tolist()]
        if 'name' in row_values or 'stakeholder' in ' '.join(row_values):
            header_row = i
            break

    if header_row is None:
        raise ValueError("Could not find header row in Carta export")

    # Re-read with correct header
    df = pd.read_excel(carta_path, sheet_name='Detailed Cap', header=header_row)

    # Parse company name from title (row 0)
    title_cell = str(raw.iloc[0, 0])
    company_name = title_cell.replace('Detailed Capitalization Table', '').strip()

    # Parse date from "As of" row (usually row 1)
    cap_table_date = None
    for i in range(min(5, len(raw))):
        cell_val = str(raw.iloc[i, 0])
        if 'As of' in cell_val:
            date_match = re.search(r'(\d{2}/\d{2}/\d{4})', cell_val)
            if date_match:
                cap_table_date = datetime.strptime(date_match.group(1), '%m/%d/%Y')
                break

    # Identify share class columns (exclude ID, Name, and summary columns)
    share_class_cols = []
    for col in df.columns:
        col_lower = col.lower()
        if any(x in col_lower for x in ['class', 'series']) and 'units' in col_lower:
            share_class_cols.append(col)

    # Identify options columns
    options_cols = [col for col in df.columns if 'option' in col.lower() or 'rsu' in col.lower()]

    # Build column index mapping for share classes (for formula references)
    # Note: header=header_row doesn't shift columns, so pandas col index = raw col index
    share_class_col_indices = {}
    for class_name in share_class_cols:
        col_idx = list(df.columns).index(class_name)
        share_class_col_indices[class_name] = col_idx

    # Build column index mapping for options columns
    options_col_indices = {}
    for opt_col in options_cols:
        col_idx = list(df.columns).index(opt_col)
        options_col_indices[opt_col] = col_idx

    # Filter to stakeholder rows only (exclude summary rows at bottom)
    # Summary rows have specific text patterns
    summary_patterns = ['total', 'outstanding', 'available', 'fully diluted', 'percentage', 'price per']

    def is_stakeholder_row(row):
        name = str(row.get('Name', '')).lower()
        return name and not any(pattern in name for pattern in summary_patterns) and name != 'nan'

    stakeholders_df = df[df.apply(is_stakeholder_row, axis=1)].copy()

    # Store original row indices for formula references
    # When header=header_row, pandas index 0 = Excel row (header_row + 2)
    # Formula: excel_row = pandas_index + header_row + 2
    stakeholders_df['_carta_row'] = stakeholders_df.index + header_row + 2

    # Calculate total shares per stakeholder
    stakeholders_df['_total_shares'] = 0
    for col in share_class_cols:
        stakeholders_df['_total_shares'] += pd.to_numeric(stakeholders_df[col], errors='coerce').fillna(0)

    # Sum options columns
    stakeholders_df['_total_options'] = 0
    for col in options_cols:
        stakeholders_df['_total_options'] += pd.to_numeric(stakeholders_df[col], errors='coerce').fillna(0)

    # Extract validation totals from summary rows
    validation = {}
    for idx, row in df.iterrows():
        name = str(row.get('Name', '')).lower()
        if 'total units outstanding' in name:
            validation['total_outstanding'] = row.get('Outstanding Units', 0)
        elif 'fully diluted' in name and 'units' in name:
            for col in share_class_cols:
                validation[f'fully_diluted_{col}'] = row.get(col, 0)

    # Extract share prices from "Price per unit" row
    price_per_unit = {}
    price_row_num = None
    for idx in range(len(raw)):
        cell_val = str(raw.iloc[idx, 1]).lower() if raw.shape[1] > 1 else ''
        if 'price per unit' in cell_val:
            price_row_num = idx + 1  # Excel 1-based row number
            # Extract prices for each share class column
            for class_name in share_class_cols:
                col_idx = share_class_col_indices[class_name]
                if col_idx < raw.shape[1]:
                    price = pd.to_numeric(raw.iloc[idx, col_idx], errors='coerce')
                    if pd.notna(price) and price > 0:
                        price_per_unit[class_name] = price
            break

    return {
        'company_name': company_name,
        'cap_table_date': cap_table_date or datetime.now(),
        'stakeholders': stakeholders_df,
        'share_class_cols': share_class_cols,
        'share_class_col_indices': share_class_col_indices,
        'options_cols': options_cols,
        'options_col_indices': options_col_indices,
        'validation': validation,
        'header_row': header_row,
        'price_per_unit': price_per_unit,
        'price_row_num': price_row_num
    }


def copy_carta_sheet_to_workbook(carta_path: str, target_wb) -> str:
    """Copy Carta 'Detailed Cap' sheet into target workbook as 'Carta Raw'."""
    carta_wb = load_workbook(carta_path)
    carta_sheet = carta_wb['Detailed Cap']

    # Create new sheet in target workbook
    sheet_name = 'Carta Raw'
    if sheet_name in target_wb.sheetnames:
        del target_wb[sheet_name]
    raw_sheet = target_wb.create_sheet(sheet_name)

    # Copy all cells including values (not formulas, since Carta exports are values)
    for row in carta_sheet.iter_rows():
        for cell in row:
            new_cell = raw_sheet[cell.coordinate]
            new_cell.value = cell.value
            # Copy cell formatting if present
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    # Copy column widths
    for col_letter, col_dim in carta_sheet.column_dimensions.items():
        raw_sheet.column_dimensions[col_letter].width = col_dim.width

    carta_wb.close()
    return sheet_name


def transform_to_template(carta_data: dict, template_path: str, output_path: str, carta_path: str) -> dict:
    """Transform Carta data into the firm's cap table template."""

    # Load template (preserves formulas)
    wb = load_workbook(template_path)
    inputs_sheet = wb['Inputs']

    # Copy Carta data sheet into this workbook
    raw_sheet_name = copy_carta_sheet_to_workbook(carta_path, wb)

    # 1. Set company name and date
    inputs_sheet['I6'] = carta_data['company_name']
    inputs_sheet['I7'] = carta_data['cap_table_date']

    # 2. Update share class headers in row 30 (columns F-I for up to 4 classes)
    share_classes = carta_data['share_class_cols']
    share_class_col_indices = carta_data['share_class_col_indices']
    class_col_mapping = {}  # Maps class name to template column number
    carta_col_mapping = {}  # Maps class name to Carta Raw column letter

    for i, class_name in enumerate(share_classes[:4]):  # Max 4 classes
        col_num = 6 + i  # Column numbers: 6, 7, 8, 9 (F, G, H, I)

        # Clean up class name for header (e.g., "Class A Units (CA)" -> "Class A")
        clean_name = re.sub(r'\s*Units.*$', '', class_name).strip()
        clean_name = re.sub(r'\s*\([^)]*\)', '', clean_name).strip()

        inputs_sheet.cell(row=30, column=col_num).value = clean_name
        class_col_mapping[class_name] = col_num

        # Get Carta Raw column letter for this share class
        carta_col_idx = share_class_col_indices[class_name]
        carta_col_mapping[class_name] = get_column_letter(carta_col_idx + 1)  # +1 for 1-based

    # Build options column letter mapping for Carta Raw
    options_col_indices = carta_data['options_col_indices']
    options_carta_cols = [get_column_letter(idx + 1) for idx in options_col_indices.values()]

    # 3. Sort stakeholders by total shares (descending) for top 9
    stakeholders = carta_data['stakeholders'].copy()
    stakeholders = stakeholders.sort_values('_total_shares', ascending=False)

    # Split into top 9 and "other"
    top_investors = stakeholders.head(9)
    other_investors = stakeholders.iloc[9:] if len(stakeholders) > 9 else pd.DataFrame()

    # 4. Populate investor rows 31-39 with FORMULAS referencing Carta Raw
    for i, (idx, row) in enumerate(top_investors.iterrows()):
        excel_row = 31 + i
        carta_row = int(row['_carta_row'])

        # Investor name - formula reference to Carta Raw
        name_formula = f"='{raw_sheet_name}'!B{carta_row}"
        inputs_sheet.cell(row=excel_row, column=4).value = name_formula

        # Share classes (columns F-I based on mapping) - formulas
        for class_name, col_num in class_col_mapping.items():
            carta_col = carta_col_mapping[class_name]
            formula = f"='{raw_sheet_name}'!{carta_col}{carta_row}"
            inputs_sheet.cell(row=excel_row, column=col_num).value = formula

        # Common shares (column P) - Carta doesn't separate, so 0
        inputs_sheet.cell(row=excel_row, column=16).value = 0

        # Options (column Q) - SUM formula for all options columns from Carta Raw
        if options_carta_cols:
            options_refs = [f"'{raw_sheet_name}'!{col}{carta_row}" for col in options_carta_cols]
            options_formula = f"={'+'.join(options_refs)}"
            inputs_sheet.cell(row=excel_row, column=17).value = options_formula
        else:
            inputs_sheet.cell(row=excel_row, column=17).value = 0

    # Clear remaining investor rows (if less than 9 investors)
    for i in range(len(top_investors), 9):
        excel_row = 31 + i
        inputs_sheet.cell(row=excel_row, column=4).value = None
        for col in range(6, 18):  # F through Q
            inputs_sheet.cell(row=excel_row, column=col).value = 0

    # 5. Populate "Other Investors" row (row 40) - SUM formulas for remaining investors
    if len(other_investors) > 0:
        inputs_sheet.cell(row=40, column=4).value = "Other Investors"

        # Build SUM formulas for each share class
        other_rows = other_investors['_carta_row'].astype(int).tolist()

        for class_name, col_num in class_col_mapping.items():
            carta_col = carta_col_mapping[class_name]
            # Create SUM formula for all "other" investor rows
            cell_refs = [f"'{raw_sheet_name}'!{carta_col}{r}" for r in other_rows]
            sum_formula = f"={'+'.join(cell_refs)}"
            inputs_sheet.cell(row=40, column=col_num).value = sum_formula

        inputs_sheet.cell(row=40, column=16).value = 0  # Common

        # Options SUM for other investors
        if options_carta_cols:
            all_opts_refs = []
            for r in other_rows:
                for col in options_carta_cols:
                    all_opts_refs.append(f"'{raw_sheet_name}'!{col}{r}")
            options_sum_formula = f"={'+'.join(all_opts_refs)}"
            inputs_sheet.cell(row=40, column=17).value = options_sum_formula
        else:
            inputs_sheet.cell(row=40, column=17).value = 0
    else:
        inputs_sheet.cell(row=40, column=4).value = "Other Investors"
        for col in range(6, 18):
            inputs_sheet.cell(row=40, column=col).value = 0

    # 6. Zero out warrants row (row 41)
    for col in range(6, 16):  # Columns F through O
        inputs_sheet.cell(row=41, column=col).value = 0

    # 7. Populate share prices if available
    # Prices go in row 10 (data row under header row 9), columns K onwards match share classes
    if carta_data.get('price_per_unit') and carta_data.get('price_row_num'):
        price_row = carta_data['price_row_num']
        # Map share class prices to columns K, L, M, N (columns 11, 12, 13, 14)
        for i, class_name in enumerate(share_classes[:4]):
            if class_name in carta_data['price_per_unit']:
                carta_col = carta_col_mapping[class_name]
                price_formula = f"='{raw_sheet_name}'!{carta_col}{price_row}"
                # Price Per Share column starts at K (column 11) for first round
                # Based on template structure, prices align with rounds in rows 10-19
                inputs_sheet.cell(row=10 + i, column=11).value = price_formula

    # 8. Save output
    wb.save(output_path)

    # 9. Generate validation summary
    total_by_class = {}
    for class_name in share_classes[:4]:
        col_total = pd.to_numeric(stakeholders[class_name], errors='coerce').fillna(0).sum()
        total_by_class[class_name] = col_total

    return {
        'output_path': output_path,
        'investors_processed': len(stakeholders),
        'top_investors': len(top_investors),
        'other_investors_count': len(other_investors),
        'share_classes_mapped': list(class_col_mapping.keys()),
        'totals_by_class': total_by_class,
        'validation': carta_data['validation'],
        'prices_found': bool(carta_data.get('price_per_unit')),
        'carta_raw_sheet': raw_sheet_name
    }


def run_transformation(carta_path: str, template_path: str, output_dir: str = None) -> dict:
    """Main entry point for the transformation."""

    carta_path = Path(carta_path)
    template_path = Path(template_path)

    if not carta_path.exists():
        raise FileNotFoundError(f"Carta export not found: {carta_path}")
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    # Parse Carta data
    carta_data = parse_carta_export(str(carta_path))

    # Generate output filename
    company_clean = re.sub(r'[^\w\s-]', '', carta_data['company_name']).strip()
    date_str = carta_data['cap_table_date'].strftime('%Y%m%d')
    output_filename = f"{company_clean}_Cap_Table_{date_str}.xlsx"

    if output_dir:
        output_path = Path(output_dir) / output_filename
    else:
        output_path = carta_path.parent / output_filename

    # Run transformation (now passing carta_path for sheet copying)
    result = transform_to_template(carta_data, str(template_path), str(output_path), str(carta_path))

    return result


# xlwings entry point
def main():
    """Called by xlwings button click."""
    import xlwings as xw

    # Get the calling workbook
    wb = xw.Book.caller()

    # File picker for Carta export
    import tkinter as tk
    from tkinter import filedialog

    root = tk.Tk()
    root.withdraw()

    carta_path = filedialog.askopenfilename(
        title="Select Carta Export",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )

    if not carta_path:
        xw.Book.caller().sheets[0].range('A1').value = "Cancelled - no file selected"
        return

    # Template path (in templates/ folder relative to project root)
    script_dir = Path(__file__).parent
    template_path = script_dir.parent / "templates" / "Cap Table Template.xlsx"

    if not template_path.exists():
        # Try looking in common locations
        template_path = filedialog.askopenfilename(
            title="Select Cap Table Template",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if not template_path:
            return

    try:
        result = run_transformation(carta_path, str(template_path))

        # Show success message
        msg = f"Success!\n\nOutput: {result['output_path']}\n"
        msg += f"Investors processed: {result['investors_processed']}\n"
        msg += f"Top investors: {result['top_investors']}\n"
        msg += f"Other investors: {result['other_investors_count']}\n"
        msg += f"Carta Raw sheet: {result['carta_raw_sheet']}\n"
        msg += f"Prices found: {result['prices_found']}"

        tk.messagebox.showinfo("Carta Transformer", msg)

        # Open the output file
        xw.Book(result['output_path'])

    except Exception as e:
        tk.messagebox.showerror("Error", str(e))


if __name__ == "__main__":
    # For testing without xlwings
    import sys
    if len(sys.argv) >= 3:
        result = run_transformation(sys.argv[1], sys.argv[2])
        print(f"Output: {result['output_path']}")
        print(f"Investors: {result['investors_processed']}")
        print(f"Carta Raw sheet: {result['carta_raw_sheet']}")
        print(f"Prices found: {result['prices_found']}")
    else:
        print("Usage: python carta_to_cap_table.py <carta_export.xlsx> <template.xlsx>")
